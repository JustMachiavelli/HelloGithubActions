from typing import List

import xlsxwriter
from openpyxl import load_workbook


def read_two_col_in_xlsx_as_dict(path: str, sheet: str, col_key: int, col_value: int):
    """
    读取xlsx中的两列值组成字典

    将前一列作key，后一列作value

    Args:
        path: xlsx路径
        sheet: sheet名称
        col_key: 作为key的列序号
        col_value: 作为value的列序号

    Returns:

    """
    return {
        row[col_key].value: row[col_value].value
        for row in load_workbook(filename=path, read_only=True)[sheet].rows
        if row[col_key].value
    }


def read_one_col_in_xlsx_as_list(path: str, sheet: str = 'Sheet1', col: int = 0):
    """
    读取xlsx中的一列值组成list

    Args:
        path: xlsx路径
        sheet: sheet名称
        col: 目标列的序号

    Returns:

    """
    return [
        row[col].value
        for row in load_workbook(filename=path, read_only=True)[sheet].rows
        if row[col].value
    ]


def read_one_row_in_xlsx_as_list(path: str, sheet: str, row: int):
    """
    读取xlsx中指定一行的内容为list

    Args:
        path: xlsx路径
        sheet: sheet名称
        row: 目标行的序号，如果是负数，表示从最后一行倒数

    Returns:
        指定一行内容list
    """
    return [
        cell.value
        for cell in list(load_workbook(filename=path, read_only=True)[sheet].rows)[row]
        if cell.value
    ]


def read_xlsx_rows_and_cols(path, sheet_name, min_row=None, max_row=None, min_col=None, max_col=None):
    """
    读xlsx的指定多少行为list

    Args:
        path: xlsx路径
        sheet_name: sheet名称
        min_row: 开始行，如果为None，则取第一行
        max_row: 结束行，如果为None，则取最后一行
        max_col: 开始列，如果为None，则取第一列
        min_col: 结束行，如果为None，则取最后一行

    Returns:
        list_rows

    Examples:
        xlsx_src = 'C:\\Users\\wanghao.su\\Desktop\\64_p8吴中三资\\DOCS\\CCB_CCIP_银企直连客户交易接口说明_信息报告_V1.9_20190822(1).xlsx'
        xlsx_write = 'C:\\Users\\wanghao.su\\Desktop\\64_p8吴中三资\\DOCS\\666.xlsx'
        list_temp = read_xlsx_exact_path_sheet(xlsx_write, '应用字典', 0, 0)
        for row in list_temp:
            print(row)
    """
    # 读取xlsx文件，对应sheet，每一行的内容
    sheet = load_workbook(filename=path)[sheet_name]
    return [
        [str(cell.value).strip() if cell.value else '' for cell in row]
        for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)
    ]


def copy_exist_sheet(list_src, sheet, align):
    for no_row in range(len(list_src)):
        # print(no_row)
        sheet.write_row(no_row, 0, list_src[no_row], align)
        sheet._write_sheet_pr()


def write_rows_to_xlsx(xlsx_path: str, rows: List[List[str]]):
    workbook = xlsxwriter.Workbook(xlsx_path)
    sheet = workbook.add_worksheet('Sheet1')
    for row_no, row in enumerate(rows):
        sheet.write_row(row_no, 0, row)
    workbook.close()
